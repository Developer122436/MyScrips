from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook(r"xlsx\regions.xlsx")
ws = wb.active

# Will show all the places that had values
cell_range = ws["A1":"C1"]
col_range = ws["A":"C"]

row_range = ws[1:5]

# iter_rows specified the max and minimum columns and rows
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)
