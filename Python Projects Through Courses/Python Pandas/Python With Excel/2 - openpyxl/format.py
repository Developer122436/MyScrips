from openpyxl.styles import (
    Font,
    colors,
    Color,
    Alignment,
    PatternFill,
    GradientFill,
    Border,
    Side,
)

from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

for i in range(1, 20):
    ws.append(range(300))

# calls the merge and center function
ws.merge_cells("A1:B5")
ws.unmerge_cells("A1:B5")
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

cell = ws["B2"]
cell.font = Font(color="FF0000", size=20, italic=True)
cell.value = "Merged Cell"
cell.alignment = Alignment(horizontal="right", vertical="bottom")
cell.fill = GradientFill(stop=("000000", "FFFFFF"))
wb.save("text.xlsx")

# name styles are object we can create that store a style that
# we can use it multiple times instead of having to type all that
# out like we did for our merged cell.
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True)

# Set the borders to be darken
bd = Side(style="thick", color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

# Set the fill color of each cell
# We will set fgColor green and red values to be the highest as they can be and no blue values - this will be yellow color.
highlight.fill = PatternFill("solid", fgColor="FFFF00")

count = 0
# in iter_cols , Column is stored as a list
# Because of that we use the counter to index the list
# for the rows we want, every time column moves over one
# our counter should add one so we move one row down.
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count = count + 1
wb.save("highlight.xlsx")
